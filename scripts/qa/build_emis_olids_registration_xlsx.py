"""Build an .xlsx summary of EMIS vs OLIDS practice registration tracking.

Pulls from int_emis_olids_practice_registration_comparison (built in DEV by
default), joins to dim_practice and dim_practice_neighbourhood for PCN and
neighbourhood roll-ups, and writes a multi-tab workbook to the caller's
Downloads folder.

Usage:
    python scripts/qa/build_emis_olids_registration_xlsx.py
    python scripts/qa/build_emis_olids_registration_xlsx.py --target prod
    python scripts/qa/build_emis_olids_registration_xlsx.py --output /tmp/out.xlsx

Requires the snowflake-connector-python package and a connections.toml entry
for the --connection name (default: data-platform-manager).
"""

from __future__ import annotations

import argparse
import pathlib
import sys
from datetime import date

import snowflake.connector
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

TARGET_DB_PREFIX = {
    "dev": "DEV__",
    "prod": "",
}

# Explanatory notes for practices where the headline % is misleading.
NOTES: dict[str, str] = {
    "F83007": (
        "Merged into Islington Central Medical Centre (F83010). EMIS list still "
        "reflects pre-merge codes; combined variance across both is -0.5%."
    ),
    "F83010": (
        "Now serves ex-Roman Way (F83007) patients post-merge. EMIS list yet to "
        "combine the two; combined variance across both is -0.5%."
    ),
    "Y03103": (
        "Medicus Select Care — specialised provider (e.g. violent patients). "
        "On EMIS software but outside the EMIS enterprise agreement, so absent "
        "from the EMIS list-size extract."
    ),
}

# Styles
HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HFILL = PatternFill("solid", start_color="1F4E78")
TITLE_FONT = Font(name="Arial", bold=True, size=14)
SUBTITLE_FONT = Font(name="Arial", italic=True, size=10)
LABEL_FONT = Font(name="Arial", bold=True)
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTRE_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
PASS_FILL = PatternFill("solid", start_color="E2EFDA")
WARN_FILL = PatternFill("solid", start_color="FFF2CC")
FAIL_FILL = PatternFill("solid", start_color="FCE4D6")
MISSING_FILL = PatternFill("solid", start_color="D9E1F2")


def category_fill(category: str | None) -> PatternFill | None:
    if not category:
        return None
    if category == "Meets Criteria":
        return PASS_FILL
    if category == "Missing Data":
        return MISSING_FILL
    if category.startswith("20%"):
        return FAIL_FILL
    if "5-20%" in category or "2-5%" in category:
        return WARN_FILL
    return None


def fetch(conn, sql: str) -> list[dict]:
    cur = conn.cursor(snowflake.connector.DictCursor)
    try:
        cur.execute(sql)
        return cur.fetchall()
    finally:
        cur.close()


def write_header_row(ws, row: int, cols: list[str]) -> None:
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = HFONT
        cell.fill = HFILL
        cell.alignment = CENTRE_WRAP
        cell.border = BORDER


def apply_widths(ws, widths: list[tuple[int, int]]) -> None:
    for col, w in widths:
        ws.column_dimensions[chr(ord("A") + col - 1)].width = w


# ---------- Sheet builders ----------

def build_summary_sheet(wb: Workbook, practices: list[dict], extract_date) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "EMIS vs OLIDS — registration tracking summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A2"] = (
        f"EMIS extract: {extract_date}  |  "
        f"All NCL practices in OLIDS + EMIS  |  Generated {date.today()}"
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    # Top-line numbers
    total_emis = sum(p["EMIS_LIST_SIZE"] or 0 for p in practices)
    total_olids = sum(p["OLIDS_REGULAR_COUNT"] or 0 for p in practices)
    net_diff = total_olids - total_emis
    net_pct = net_diff / total_emis * 100 if total_emis else 0
    n_pass = sum(1 for p in practices if p["VARIANCE_CATEGORY"] == "Meets Criteria")
    n_within_5 = sum(
        1 for p in practices
        if p["ABSOLUTE_PERCENT_DIFFERENCE"] is not None
        and p["ABSOLUTE_PERCENT_DIFFERENCE"] < 5
    )
    n_total = len(practices)

    ws["A4"] = "Headline"
    ws["A4"].font = LABEL_FONT
    ws["B4"] = (
        f"OLIDS tracks EMIS within {abs(net_pct):.2f}% at NCL scale "
        f"({net_diff:+,} patients across {n_total:,} practices, {total_emis:,} EMIS total).  "
        f"{n_pass}/{n_total} practices meet acceptance criteria (<2% or <5 persons).  "
        f"{n_within_5}/{n_total} are within 5% of EMIS — only a handful of outliers, "
        f"mostly upstream backfill gaps or known data exceptions."
    )
    ws["B4"].alignment = WRAP_TOP
    ws.merge_cells("B4:F4")
    ws.row_dimensions[4].height = 60

    # Category breakdown
    write_header_row(ws, 6, ["Category", "Practices", "EMIS list size", "OLIDS count", "Net diff"])
    cats_order = ["Meets Criteria", "2-5% Variance", "5-20% Variance", "20%+ Variance", "Missing Data"]
    by_cat = {c: {"n": 0, "emis": 0, "olids": 0, "diff": 0} for c in cats_order}
    for p in practices:
        c = p["VARIANCE_CATEGORY"] or "Missing Data"
        if c not in by_cat:
            by_cat[c] = {"n": 0, "emis": 0, "olids": 0, "diff": 0}
        by_cat[c]["n"] += 1
        by_cat[c]["emis"] += p["EMIS_LIST_SIZE"] or 0
        by_cat[c]["olids"] += p["OLIDS_REGULAR_COUNT"] or 0
        by_cat[c]["diff"] += p["DIFFERENCE"] or 0

    row = 7
    for c in cats_order:
        v = by_cat[c]
        cells = [c, v["n"], v["emis"], v["olids"], v["diff"]]
        for i, val in enumerate(cells, 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.border = BORDER
            if i in (3, 4):
                cell.number_format = "#,##0"
            if i == 5:
                cell.number_format = "+#,##0;-#,##0;0"
            fill = category_fill(c)
            if fill:
                cell.fill = fill
        row += 1

    # Total row
    cells = ["TOTAL", n_total, total_emis, total_olids, net_diff]
    for i, val in enumerate(cells, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.border = BORDER
        cell.font = LABEL_FONT
        if i in (3, 4):
            cell.number_format = "#,##0"
        if i == 5:
            cell.number_format = "+#,##0;-#,##0;0"

    apply_widths(ws, [(1, 20), (2, 12), (3, 16), (4, 16), (5, 14), (6, 14)])


def build_practice_sheet(wb: Workbook, practices: list[dict]) -> None:
    """Practice-level table, sorted worst → best by |% variance|."""
    ws = wb.create_sheet("Practices (worst → best)")

    cols = [
        "Practice code", "Practice name", "Borough", "PCN",
        "EMIS list", "OLIDS count", "Difference", "% variance",
        "Abs % variance", "Category", "Notes",
    ]
    write_header_row(ws, 1, cols)

    # Sort: nulls first (Missing Data), then by |%| descending, then practice code
    def sort_key(p):
        ap = p["ABSOLUTE_PERCENT_DIFFERENCE"]
        return (ap is not None, -(ap or 0), p["PRACTICE_CODE"])

    sorted_practices = sorted(practices, key=sort_key)

    for ridx, p in enumerate(sorted_practices, start=2):
        cat = p["VARIANCE_CATEGORY"]
        fill = category_fill(cat)
        values = [
            p["PRACTICE_CODE"], p["PRACTICE_NAME"], p["BOROUGH"], p["PCN_NAME"],
            p["EMIS_LIST_SIZE"], p["OLIDS_REGULAR_COUNT"], p["DIFFERENCE"],
            (p["PERCENT_DIFFERENCE"] / 100) if p["PERCENT_DIFFERENCE"] is not None else None,
            (p["ABSOLUTE_PERCENT_DIFFERENCE"] / 100) if p["ABSOLUTE_PERCENT_DIFFERENCE"] is not None else None,
            cat,
            NOTES.get(p["PRACTICE_CODE"], ""),
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ridx, column=ci, value=val)
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if ci in (5, 6):
                cell.number_format = "#,##0"
            if ci == 7:
                cell.number_format = "+#,##0;-#,##0;0"
            if ci in (8, 9):
                cell.number_format = "+0.00%;-0.00%;0.00%"
            if ci == 11:
                cell.alignment = WRAP_TOP

    apply_widths(ws, [(1, 14), (2, 38), (3, 13), (4, 32), (5, 12), (6, 12),
                      (7, 12), (8, 12), (9, 12), (10, 16), (11, 70)])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(sorted_practices) + 1}"


def build_rollup_sheet(
    wb: Workbook,
    title: str,
    sheet_name: str,
    practices: list[dict],
    group_field: str,
    group_label: str,
) -> None:
    """Generic group-by rollup tab (borough / PCN / neighbourhood)."""
    ws = wb.create_sheet(sheet_name)

    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")

    cols = [group_label, "Practices", "Passing", "Pass rate", "EMIS list", "OLIDS count", "Net diff", "Net %"]
    write_header_row(ws, 3, cols)

    groups: dict[str, dict] = {}
    for p in practices:
        key = p[group_field] or "(unknown)"
        g = groups.setdefault(key, {"n": 0, "passing": 0, "emis": 0, "olids": 0, "diff": 0})
        g["n"] += 1
        if p["VARIANCE_CATEGORY"] == "Meets Criteria":
            g["passing"] += 1
        g["emis"] += p["EMIS_LIST_SIZE"] or 0
        g["olids"] += p["OLIDS_REGULAR_COUNT"] or 0
        g["diff"] += p["DIFFERENCE"] or 0

    rows = []
    for key, g in groups.items():
        net_pct = g["diff"] / g["emis"] if g["emis"] else None
        pass_rate = g["passing"] / g["n"] if g["n"] else None
        rows.append([key, g["n"], g["passing"], pass_rate, g["emis"], g["olids"], g["diff"], net_pct])

    # Sort by absolute net % descending (worst on top), nulls last
    rows.sort(key=lambda r: (r[7] is None, -abs(r[7] or 0)))

    # Column mapping: 1=group, 2=practices, 3=passing, 4=pass_rate, 5=emis, 6=olids, 7=diff, 8=net%
    for ridx, row_vals in enumerate(rows, start=4):
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ridx, column=ci, value=val)
            cell.border = BORDER
            if ci == 4:
                cell.number_format = "0%"
            if ci in (5, 6):
                cell.number_format = "#,##0"
            if ci == 7:
                cell.number_format = "+#,##0;-#,##0;0"
            if ci == 8:
                cell.number_format = "+0.00%;-0.00%;0.00%"

    # Total row
    total_n = sum(r[1] for r in rows)
    total_pass = sum(r[2] for r in rows)
    total_emis = sum(r[5] for r in rows)
    total_diff = sum(r[7] for r in rows)
    totals = [
        "TOTAL",
        total_n,
        total_pass,
        total_pass / total_n if total_n else None,
        total_emis,
        sum(r[6] for r in rows),
        total_diff,
        total_diff / total_emis if total_emis else None,
    ]
    total_row = len(rows) + 4
    for ci, val in enumerate(totals, 1):
        cell = ws.cell(row=total_row, column=ci, value=val)
        cell.border = BORDER
        cell.font = LABEL_FONT
        if ci == 4:
            cell.number_format = "0%"
        if ci in (5, 6):
            cell.number_format = "#,##0"
        if ci == 7:
            cell.number_format = "+#,##0;-#,##0;0"
        if ci == 8:
            cell.number_format = "+0.00%;-0.00%;0.00%"

    apply_widths(ws, [(1, 40), (2, 11), (3, 10), (4, 10), (5, 14), (6, 14), (7, 12), (8, 10)])
    ws.freeze_panes = "A4"


# ---------- Main ----------

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--connection", default="data-platform-manager",
                        help="Snowflake connection_name from connections.toml")
    parser.add_argument("--target", choices=["dev", "prod"], default="dev",
                        help="Which environment to read from (default: dev)")
    parser.add_argument("--output", type=pathlib.Path,
                        default=pathlib.Path.home() / "Downloads" / "olids_emis_registration_summary.xlsx",
                        help="Output path (default: ~/Downloads/olids_emis_registration_summary.xlsx)")
    args = parser.parse_args()

    prefix = TARGET_DB_PREFIX[args.target]
    comp_table = f"{prefix}MODELLING.OLIDS_UTILITIES.INT_EMIS_OLIDS_PRACTICE_REGISTRATION_COMPARISON"
    dim_practice = f"{prefix}REPORTING.OLIDS_ORGANISATION.DIM_PRACTICE"
    dim_neigh = f"{prefix}REPORTING.OLIDS_ORGANISATION.DIM_PRACTICE_NEIGHBOURHOOD"

    print(f"Connecting via {args.connection} (target: {args.target})...")
    conn = snowflake.connector.connect(connection_name=args.connection, authenticator="externalbrowser")
    try:
        # Practice-level data, joined to dim_practice + dim_practice_neighbourhood
        practices = fetch(conn, f"""
            SELECT
                c.practice_code,
                c.practice_name,
                c.borough,
                dp.pcn_name_with_borough AS pcn_name,
                dpn.neighbourhood_registered,
                c.emis_list_size,
                c.olids_regular_count,
                c.difference,
                c.percent_difference,
                c.absolute_percent_difference,
                c.variance_category,
                c.extract_date
            FROM {comp_table} c
            LEFT JOIN {dim_practice} dp ON c.practice_code = dp.practice_code
            LEFT JOIN {dim_neigh} dpn ON c.practice_code = dpn.practice_code
        """)
        print(f"Loaded {len(practices)} practices.")
        if not practices:
            print("No rows returned — has int_emis_olids_practice_registration_comparison been built?",
                  file=sys.stderr)
            return 1

        extract_date = next((p["EXTRACT_DATE"] for p in practices if p["EXTRACT_DATE"]), None)

        # Build workbook
        wb = Workbook()
        build_summary_sheet(wb, practices, extract_date)
        build_practice_sheet(wb, practices)
        build_rollup_sheet(wb, "Borough rollup", "Borough", practices, "BOROUGH", "Borough")
        build_rollup_sheet(wb, "PCN rollup", "PCN", practices, "PCN_NAME", "PCN")
        build_rollup_sheet(wb, "Neighbourhood rollup", "Neighbourhood",
                           practices, "NEIGHBOURHOOD_REGISTERED", "Neighbourhood")

        args.output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(args.output)
        print(f"Saved: {args.output}")
    finally:
        conn.close()

    return 0


if __name__ == "__main__":
    sys.exit(main())
